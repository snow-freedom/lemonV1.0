import requests
import openpyxl

# #读取测试用例
def read_data(filename,sheetname):
    wb = openpyxl.load_workbook(filename)
    sh = wb[sheetname]
    max_row = sh.max_row
    case_list = []
    for i in range(2,max_row+1):
        dict1 = dict(
        case_id = sh.cell(row=i,column=1).value,
        url = sh.cell(row=i,column=5).value,
        data = sh.cell(row=i,column=6).value,
        expect = sh.cell(row=i,column=7).value
        )
        case_list.append(dict1)
    return case_list
case = read_data('C:\\pathon_workspace\\test_data\\test_case_api.xlsx', 'login')
# print(case)



#发送接口测试
def api_fun(url,data):    #定义函数名称,并定义参数
    # url_login = 'http://8.129.91.152:8766/futureloan/member/login'#请求地址
    # data_login = {"mobile_phone": "18791767359","pwd": "lemon666"}#请求正文
    headers = {'X-Lemonban-Media-Type':'lemonban.v2','Content-Type':'application/json'}#封装函数中的请求头必须是headers = 格式

    result = requests.post(url=url,json=data,headers=headers).json()
    return result
#     # print(result)

#执行接口测试
def wirte_value(filename,sheetname,row,column,final_result):
    wb = openpyxl.load_workbook(filename)
    sh = wb[sheetname]
    sh.cell(row=row,column=column).value = final_result
    wb.save(filename)
# wirte_value('test_case_api.xlsx','register',5,8,'pass')

#接口测试实战,分装成函数
def execute_fun(filename,sheetname):
    cases = read_data(filename,sheetname)
    # print(cases)
    for case in cases:
        case_id = case['case_id']
        url = case['url']
        data = eval(case['data'])
        # print(case_id,url,data)

        #获取期望结果code、msg
        expect = eval(case['expect'])
        expect_code = expect['code']
        expect_msg = expect['msg']
        print('期望的code:{}'.format(expect_code),'期望的msg:{}'.format(expect_msg))

        # 执行测试用例
        real_results = api_fun(url=url,data=data)
        # print(real_results)
        real_code = real_results['code']
        real_msg = real_results['msg']
        print('实际的code:{}'.format(real_code), '实际的msg:{}'.format(real_msg))
        print('#'*50)

        #断言
        if real_code == expect_code and real_msg == expect_msg:
            print('这条用例执行通过')
            final_result = 'passed'
        else:
            print('这条用例不通过')
            final_result = 'fialed'
        #写入excel表
        wirte_value(filename,sheetname,case_id+1,8,final_result)
# execute_fun('../test_data/test_case_api.xlsx', 'login')

